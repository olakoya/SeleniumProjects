'''
Data Driven Testing
--------------------

Multiple sets of Data is required to test ---> Functionality (Test case) in the area of validation and this is known as ---> Data Driven Testing

Data Driven Testing are performed by Excel Files which is known as Database

To handle Excel Files we use Pandas or Openpyxl

Understanding Structure of Excel File
----------------------------------------
File --> load Workbook ---> to get Sheets ---> Rows and Columns ---> Cell(r,c)

Operations on Excel Files
-------------------------
This contains count of rows and columns
Reading data from all rows and columns, also specific column data
Write data into multiple cells

Read the Data from Cell
-------------------------
To read data, it contains sheet and cell of which will return cell i.e. data
= sheet.cell(r,c).value

How to write a Data from Cell?
-------------------------------
sheet.cell(r,c).value = data ---> save

when we are writing the data we should always close the file on which we are working

Data Driven Testing
------------------
Duplication of the code
Maintainability

Utility ==> a file contains reusable functions

To install packages
-------------------
pip install -r requirements.txt
pip install openpyxl

Range functions generate --> sequence of numbers
Range(10) --> 0 to 9
Range(11) --> 0 to 10
Range(1,11) --> 1 to 10

Press --> ctrl + s --> to save the data

100 Test Cases equals --> 100 Automation code + 100 same excel operations

Possible Problems are:
---------------------
Duplication issues
Maintenance issues

Possible Answer is
-------------------
Creating Utilities file

Utility files contain --> reusable methods and functions

E.g
'''

import openpyxl
import os

file = r"/Users/olakoya/Desktop/Selenium/Day24/caldata.xlsx"

# try:
#     workbook = openpyxl.load_workbook(file)
#     print("Available sheets:", workbook.sheetnames)  # Print all sheet names
#
#     # Select the first available sheet instead of assuming "sheet1" exists
#     sheet_name = workbook.sheetnames[0]
#     sheet = workbook[sheet_name]
#     print(f"Using sheet: {sheet_name}")
#
# except Exception as e:
#     print("Error:", e)
'''
Output is
Available sheets: ['Sheet1']
Using sheet: Sheet1
'''

# Row and Colum Count
workbook = openpyxl.load_workbook(file)
# sheet = workbook["Sheet1"]
# sheet = workbook["Sheet2"]

# Reading the Data

# print(sheet.max_row)
# print(sheet.max_column)
# '''
# Output is
# 7
# 7
# '''
#
# # Reading a specific Cell Data
# # data = sheet.cell(4, 4).value
# # print(data)
# '''
# Output is
# Compounded Annually
# '''
#
# # data = sheet.cell(2, 6).value
# # print(data)
# '''
# Output is
# Pass
# '''
#
# # Reading all rows and columns Data
# data = sheet.cell(2, 6).value
# print(data)
#
# for r in range(1,sheet.max_row+1): # rows
#     for c in range(1,sheet.max_column+1): # columns
#         data = sheet.cell(r,c).value
#         print(data, end = " ")
#     print()

'''
Output is
Initial Deposit Am
Interest Rate
Length (Months)
Compounding
Total
Expected
500.0
2.0
24.0
Compounded Monthly
$520.39
Pass
1000.0
4.0
24.0
Compounded Monthly
$1,083.14
Pass
1500.0
2.0
48.0
Compounded Annually
$1,623.65
Pass
3000.0
4.0
24.0
Compounded Annually
$3,244.89
Pass
5000.0
2.0
48.0
Compounded Monthly
$0
Fail
None
None
None
None
None
None

Output for code line 108 print(data, end = " ")
Initial Deposit Am Interest Rate Length (Months) Compounding Total Expected 500.0 2.0 24.0 Compounded Monthly $520.39 
Pass 1000.0 4.0 24.0 Compounded Monthly $1,083.14 Pass 1500.0 2.0 48.0 Compounded Annually $1,623.65 Pass 3000.0 4.0 24.0 
Compounded Annually $3,244.89 Pass 5000.0 2.0 48.0 Compounded Monthly $0 Fail None None None None None None 


Output after adding line 190 print()
Initial Deposit Am Interest Rate Length (Months) Compounding Total Expected None 
500.0 2.0 24.0 Compounded Monthly $520.39 Pass None 
1000.0 4.0 24.0 Compounded Monthly $1,083.14 Pass None 
1500.0 2.0 48.0 Compounded Annually $1,623.65 Pass None 
3000.0 4.0 24.0 Compounded Annually $3,244.89 Pass None 
5000.0 2.0 48.0 Compounded Monthly $0 Fail None 
None None None None None None None 
'''

# Writing the Data
# sheet.cell(3,3).value = "Ola"
#
# workbook.save(file)

# Check if "Sheet2" exists; if not, create it (Codes from ChatGpt when the above codes was given me error output)
if "Sheet2" not in workbook.sheetnames:
    sheet2 = workbook.create_sheet("Sheet2")  # Create new sheet
    print("Sheet2 created successfully.")
else:
    sheet2 = workbook["Sheet2"]  # Select existing Sheet2
    print("Sheet2 already exists.")

# Write data to a specific row and column (e.g., row 3, column 2)
sheet2.cell(row=1, column=1, value="Hello from Sheet2!")
sheet2.cell(row=1, column=2, value="Hello from Sheet2!")
sheet2.cell(row=1, column=3, value="Hello from Sheet2!")
sheet2.cell(row=2, column=1, value="Hello from Ola!")
sheet2.cell(row=2, column=2, value="Hello from Ola!")
sheet2.cell(row=2, column=3, value="Hello from Ola!")
sheet2.cell(row=3, column=1, value="Hello from me!")
sheet2.cell(row=3, column=2, value="Hello from me!")
sheet2.cell(row=3, column=3, value="Hello from me!")

# Save the changes
workbook.save(file)
print("Data written and workbook saved successfully.")